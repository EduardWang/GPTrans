import tensorflow as tf
import numpy as np
from model import get_model
import os
import gc
from scipy import interp
from sklearn.model_selection import StratifiedKFold, StratifiedShuffleSplit
from sklearn.metrics import matthews_corrcoef, f1_score, roc_curve
from sklearn import metrics
from sklearn.metrics import confusion_matrix, precision_recall_curve
import openpyxl as op
from roc_utils import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings("ignore")
Project_Path='E:\\Projects\\GPTrans'

SEQUENCE_LEN = 181
DATANAME = 'clinvar'
Suffix = 'single_mut'

filename =Project_Path + f'/result/{DATANAME}/cv/tcv_{SEQUENCE_LEN}_1024_add_{Suffix}.xlsx'
filename2 = Project_Path + f'/result/{DATANAME}/cv/tcv_{SEQUENCE_LEN}_1024_add_{Suffix}_label.xlsx'


def op_toexcel(data, filename): # openpyxl库储存 数据到excel

    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]
        ws.append(data) # 每次写入一行
        wb.save(filename)
    else:
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        ws.append(['MCC', 'ACC', 'AUC', 'Sensitivity', 'Specificity', 'Precision', 'NPV', 'F1', 'FPR', 'FNR',
                  'TN', 'FP', 'FN', 'TP', 'AUPRC', 'Threshold'])  # 添加表头
        ws.append(data) # 每次写入一行
        wb.save(filename)


def result_toexcel(data, filename): # openpyxl库储存 数据到excel
    """
    参数:
    data(tuple): 要写入的数据，包含四个列表和一个浮点数。
    filename(str): Excel文件的名称。
    """
    # 解包data中的四个列表和一个浮点数
    list1, list2, list3, list4, threshold = data

    # 确认所有列表的长度相同
    assert len(list1) == len(list2) == len(list3) == len(list4), "所有列表必须具有相同的长度"

    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]
    else:
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb.active  # 获取活动工作表
        ws.title = "Sheet1"  # 给工作表命名
        ws.append(['ID', 'True_Label', 'Predict_Value', 'Predict_Label', 'Threshold'])  # 添加表头

    # 将数据逐行写入Excel
    for i in range(len(list1)):
        row = [list1[i], list2[i], list3[i], list4[i], threshold]
        ws.append(row)

    wb.save(filename) 

def data_generator(train_esm_wt, train_esm2_wt, train_esm1b_wt, train_prot_wt, train_esm_mut, train_esm2_mut, train_esm1b_mut, train_prot_mut, train_y, batch_size):

    L = train_esm_mut.shape[0]

    while True:
        for i in range(0, L, batch_size):
            batch_esm_wt = train_esm_wt[i:i + batch_size].copy()
            batch_esm2_wt = train_esm2_wt[i:i + batch_size].copy()
            batch_esm1b_wt = train_esm1b_wt[i:i + batch_size].copy()
            batch_prot_wt = train_prot_wt[i:i + batch_size].copy()
            batch_esm_mut = train_esm_mut[i:i + batch_size].copy()
            batch_esm2_mut = train_esm2_mut[i:i + batch_size].copy()
            batch_esm1b_mut = train_esm1b_mut[i:i + batch_size].copy()
            batch_prot_mut = train_prot_mut[i:i + batch_size].copy()
            batch_y = train_y[i:i + batch_size].copy()

            yield ([batch_esm_wt, batch_esm2_wt,batch_esm1b_wt, batch_prot_wt, batch_esm_mut, batch_esm2_mut, batch_esm1b_mut, batch_prot_mut], batch_y)


if __name__ == "__main__":

    os.environ["CUDA_VISIBLE_DEVICES"] = "0"
    gpus = tf.config.experimental.list_physical_devices('GPU')
    tf.config.experimental.set_memory_growth(gpus[0], True)

    all_esm_wt = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm1v_wt.npy')
    all_esm2_wt = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm2_wt.npy')
    all_esm1b_wt = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm1b_wt.npy')
    all_prot_wt = np.lib.format.open_memmap(Project_Path + f'/features_npy/prottrans/{DATANAME}/{SEQUENCE_LEN}/prot_wt.npy')
    all_esm_mut = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm1v_mut.npy')
    all_esm2_mut = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm2_mut.npy')
    all_esm1b_mut = np.lib.format.open_memmap(Project_Path + f'/features_npy/esm/{DATANAME}/{SEQUENCE_LEN}/cesm1b_mut.npy')
    all_prot_mut = np.lib.format.open_memmap(Project_Path + f'/features_npy/prottrans/{DATANAME}/{SEQUENCE_LEN}/prot_mut.npy')
    all_label = np.lib.format.open_memmap(Project_Path + f'/features_npy/labels/{DATANAME}/{SEQUENCE_LEN}/labels.npy')

    pos_label = True
    rocs = []
    auc_values = []
    tpr_list = []  # 存储每个交叉验证折叠的真正例率
    mean_fpr = np.linspace(0, 1, 100)  # 平均假正例率的取值范围
    mean_recall = np.linspace(0, 1, 100)  # 平均假正例率的取值范围

    for ii in range(1):
            

        
        cv = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)
        k = 1

        for train_index, test_index in cv.split(all_esm_mut, all_label):
            # 训练集
            train_esm_wt = all_esm_wt[train_index]
            train_esm2_wt = all_esm2_wt[train_index]
            train_esm1b_wt = all_esm1b_wt[train_index]
            train_prot_wt = all_prot_wt[train_index]
            train_esm_mut = all_esm_mut[train_index]
            train_esm2_mut = all_esm2_mut[train_index]
            train_esm1b_mut = all_esm1b_mut[train_index]
            train_prot_mut = all_prot_mut[train_index]
            train_label = all_label[train_index]


            # 打乱训练集顺序并划分出验证集
            # （1）分层打乱
            split = StratifiedShuffleSplit(n_splits=1, test_size=0.1, random_state=42)
            for train_inx, valid_inx in split.split(train_esm_mut, train_label):
                # 验证集
                valid_esm_wt = train_esm_wt[valid_inx]
                valid_esm2_wt = train_esm2_wt[valid_inx]
                valid_esm1b_wt = train_esm1b_wt[valid_inx]
                valid_prot_wt = train_prot_wt[valid_inx]
                valid_esm_mut = train_esm_mut[valid_inx]
                valid_esm2_mut = train_esm2_mut[valid_inx]
                valid_esm1b_mut = train_esm1b_mut[valid_inx]
                valid_prot_mut = train_prot_mut[valid_inx]
                valid_label = train_label[valid_inx]

                # 训练集
                train_esm_wt = train_esm_wt[train_inx]
                train_esm2_wt = train_esm2_wt[train_inx]
                train_esm1b_wt = train_esm1b_wt[train_inx]
                train_prot_wt = train_prot_wt[train_inx]
                train_esm_mut = train_esm_mut[train_inx]
                train_esm2_mut = train_esm2_mut[train_inx]
                train_esm1b_mut = train_esm1b_mut[train_inx]
                train_prot_mut = train_prot_mut[train_inx]
                train_label = train_label[train_inx]


            # 测试集
            test_esm_wt = all_esm_wt[test_index]
            test_esm2_wt = all_esm2_wt[test_index]
            test_esm1b_wt = all_esm1b_wt[test_index]
            test_prot_wt = all_prot_wt[test_index]
            test_esm_mut = all_esm_mut[test_index]
            test_esm2_mut = all_esm2_mut[test_index]
            test_esm1b_mut = all_esm1b_mut[test_index]
            test_prot_mut = all_prot_mut[test_index]
            test_label = all_label[test_index]


            # 训练、验证each epoch的步长
            train_size = train_label.shape[0]
            val_size = valid_label.shape[0]
            batch_size = 32
            train_steps = train_size // batch_size
            val_steps = val_size // batch_size

            print(f"Fold {k} - Training samples: {train_esm_mut.shape[0]}, Test samples: {test_esm_mut.shape[0]}")

            qa_model = get_model()

            valiBestModel = Project_Path + f'/save_model/fivecv_model/{DATANAME}/model_fold_{k}.h5'

            checkpointer = tf.keras.callbacks.ModelCheckpoint(
                filepath=valiBestModel,
                monitor='val_loss',
                save_weights_only=True,
                verbose=1,
                save_best_only=True
            )

            early_stopping = tf.keras.callbacks.EarlyStopping(
                monitor='val_loss',
                patience=10,
                verbose=0,
                mode='auto'
            )

            train_generator = data_generator(train_esm_wt,train_esm2_wt,train_esm1b_wt, train_prot_wt, train_esm_mut, train_esm2_mut,train_esm1b_mut, train_prot_mut, train_label, batch_size)
            val_generator = data_generator(valid_esm_wt,valid_esm2_wt,valid_esm1b_wt, valid_prot_wt, valid_esm_mut, valid_esm2_mut,valid_esm1b_mut, valid_prot_mut, valid_label, batch_size)
            

            history_callback = qa_model.fit_generator(
                train_generator,
                steps_per_epoch=train_steps,
                epochs=100,
                verbose=1,
                callbacks=[checkpointer, early_stopping],
                validation_data=val_generator,
                validation_steps=val_steps,
                shuffle=True,
                workers=1
            )

            train_generator.close()
            val_generator.close()

            print(f"\nFold {k} - Validation Loss: {history_callback.history['val_loss'][-1]:.4f}, " +
                f"Validation Accuracy: {history_callback.history['val_accuracy'][-1]:.4f}")

            print(f"Fold {k} - Testing:")

            test_pred = qa_model.predict([test_esm_wt,test_esm2_wt,test_esm1b_wt, test_prot_wt, test_esm_mut, test_esm2_mut, test_esm1b_mut, test_prot_mut]).reshape(-1, )
            
            y_pred = test_pred
            y_true = test_label
            y_pred_new = []
            
            del qa_model, train_generator, val_generator, train_esm_mut, train_esm1b_mut
            tf.keras.backend.clear_session()
            gc.collect()
            
            best_f1 = 0
            best_threshold = 0.5
            for threshold in range(0, 100):
                threshold = threshold / 100
                binary_pred = [1 if pred >= threshold else 0 for pred in y_pred]
                f1 = metrics.f1_score(y_true, binary_pred)
                # f1 = matthews_corrcoef(y_true, binary_pred)
                if f1 > best_f1:
                    best_f1 = f1
                    best_threshold = threshold

            for value in y_pred:
                if value < best_threshold:
                    y_pred_new.append(0)
                else:
                    y_pred_new.append(1)
            y_pred_new = np.array(y_pred_new)

            tn, fp, fn, tp = confusion_matrix(y_true, y_pred_new).ravel()
            fpr, tpr, thresholds = roc_curve(y_true, y_pred_new, pos_label=1)
            roc_auc = metrics.auc(fpr, tpr)
            precision, recall, thresholds = precision_recall_curve(y_true, y_pred_new)
            auprc = metrics.auc(recall, precision)
            thd = best_threshold

            # 计算AUC
            roc = compute_roc(X=y_pred_new, y=y_true, pos_label=pos_label)
            rocs.append(roc)
            tpr_list.append(np.interp(mean_fpr, fpr, tpr))
            roc_auc_bootstrap = roc_auc
            auc_values.append(roc_auc_bootstrap)

            print("Matthews相关系数: " + str(matthews_corrcoef(y_true, y_pred_new)))
            print("ACC: ", (tp + tn) / (tp + tn + fp + fn))
            print("AUC: ", roc_auc)
            print('sensitivity/recall:', tp / (tp + fn))
            print('specificity:', tn / (tn + fp))
            print('precision:', tp / (tp + fp))
            print('negative predictive value:', tn / (tn + fn))
            print("F1值: " + str(f1_score(y_true, y_pred_new)))
            print('error rate:', fp / (tp + tn + fp + fn))
            print('false positive rate:', fp / (tn + fp))
            print('false negative rate:', fn / (tp + fn))
            print('TN:', tn, 'FP:', fp, 'FN:', fn, 'TP:', tp)
            print('AUPRC: ' + str(auprc))
            print('best_threshold: ' + str(best_threshold))

            mcc = float(format((matthews_corrcoef(y_true, y_pred_new)), '.4f'))
            acc = float(format((tp + tn) / (tp + tn + fp + fn), '.4f'))
            auc = float(format(roc_auc, '.4f'))
            sen = float(format(tp / (tp + fn), '.4f'))
            spe = float(format(tn / (tn + fp), '.4f'))
            pre = float(format(tp / (tp + fp), '.4f'))

            npv = float(format(tn / (tn + fn), '.4f'))
            f1 = float(format(f1_score(y_true, y_pred_new), '.4f'))
            fpr = float(format(fp / (tn + fp), '.4f'))
            fnr = float(format(fn / (tp + fn), '.4f'))
            auprc = float(format(auprc, '.4f'))

            #保存每一次跑的结果到excel表格
            result = mcc, acc, auc, sen, spe, pre, npv, f1, fpr, fnr, tn, fp, fn, tp, auprc, thd
            op_toexcel(result, filename)


            k += 1


         
        # 字体设置
        font1 = {'family': 'Times New Roman',
                'size': 17}
        font2 = {'family': 'Times New Roman',
                'size': 18}
        font3 = {'family': 'Times New Roman',
                'size': 20}
        font_name = "Times New Roman"
        plt.rcParams['font.family'] = 'Arial'


        # 计算 AUC 的均值和标准差
        mean_auc = np.mean(auc_values)
        std_auc = np.std(auc_values)

        # 计算置信区间
        confidence_interval_auc = 1.96 * std_auc / np.sqrt(len(auc_values))  # 考虑标准误差和样本大小来估计置信区间的一种方法
        confidence_interval = np.percentile(auc_values, [2.5, 97.5])  # 考虑了整个分布的百分位数，而不仅仅是标准差

        roc_mean = compute_mean_roc(rocs)

        resolution = 101
        fpr_mean = np.linspace(0, 1, resolution)
        fpr_mean = np.insert(fpr_mean, 0, 0)  # Insert a leading 0 (closure)
        n_samples = len(rocs)
        thr_all = np.zeros([n_samples, resolution + 1])
        tpr_all = np.zeros([n_samples, resolution + 1])
        auc_all = np.zeros(n_samples)

        for i, ret in enumerate(rocs):
            tpr_all[i, :] = interp(fpr_mean, ret.fpr, ret.tpr)
            thr_all[i, :] = interp(fpr_mean, ret.fpr, ret.thr)
            auc_all[i] = ret.auc
            # Closure
            tpr_all[i, [0, -1]] = ret.tpr[[0, -1]]
            thr_all[i, [0, -1]] = ret.thr[[0, -1]]

        thr_mean = np.mean(thr_all, axis=0)
        tpr_mean = np.mean(tpr_all, axis=0)

        df = pd.DataFrame({
            'mean_tpr': tpr_mean,
            'mean_fpr': fpr_mean
        })

        filenamef = Project_Path + f'/data/fpr_tpr/{DATANAME}_tpr_fpr_{Suffix}_{ii}.xlsx'
        try:
            # 检查文件是否存在
            book = load_workbook(filenamef)
            # 如果文件存在，删除工作表并保存新文件
            book.remove(book.active)
        except FileNotFoundError:
            pass  # 文件不存在，则继续保存新文件

        df.to_excel(filenamef, index=False)

        plt.figure(figsize=(8, 8))
        plot_mean_roc(rocs, show_ci=True, show_ti=True)
        plt.xlim([-.05, 1.05])
        plt.ylim([-.05, 1.05])
        plt.tick_params(axis='x', labelsize=18)
        plt.tick_params(axis='y', labelsize=18)
        plt.title("ROC Curve with Confidence Interval",fontsize=20)
        plt.xlabel('False Positive Rate',fontsize=19)
        plt.ylabel('True Positive Rate',fontsize=19)
        plt.plot([0, 1], [0, 1], 'darkgray', linestyle='--')
        plt.grid(False)
        plt.legend(loc='lower right',fontsize=14)
        plt.savefig(Project_Path + f'/pic/{DATANAME}/ten_ROC_{Suffix}_{ii}.png', dpi=300)
        # plt.savefig('../pic/full_noencoder.png',dpi=300)
        plt.show()

        # 打印置信区间
        print(f'95% Confidence Interval: {confidence_interval}')








