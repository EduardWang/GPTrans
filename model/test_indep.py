import numpy as np
import os
from sklearn.metrics import matthews_corrcoef, f1_score, roc_curve
from sklearn import metrics
from sklearn.metrics import confusion_matrix, auc, precision_recall_curve
import openpyxl as op
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings("ignore")

def op_toexcel(data,filename): # openpyxl库储存 数据到excel

    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]

        ws.append(data) # 每次写入一行
        wb.save(filename)
    else:
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        ws.append(['MCC', 'ACC', 'AUC', 'Sensitivity', 'Specificity', 'Precision', 'NPV', 'F1', 'FPR', 'FNR',
                  'TN', 'FP', 'FN', 'TP','AUPRC','Threshold'])  # 添加表头
        ws.append(data) # 每次写入一行
        wb.save(filename)


def plot_roc_curve(y_true, y_pred):
    fpr, tpr, thresholds = metrics.roc_curve(y_true, y_pred)
    roc_auc = metrics.auc(fpr, tpr)

    plt.figure()
    lw = 2

    plt.plot(fpr, tpr, color='Red', lw=lw, label='ROC curve (area = %0.3f)' % roc_auc)
    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate', fontname='Arial')
    plt.ylabel('True Positive Rate', fontname='Arial')
    plt.tick_params(labelsize=10)
    plt.title('Receiver Operating Characteristic', fontsize=10)
    plt.legend(loc="lower right")
    plt.show()



def fcvtest(test_pred, test_label, filename):

    y_pred =test_pred
    y_true =test_label
    y_pred_new = []

    best_f1 = 0
    best_threshold = 0.5
    for threshold in range(10, 1000):
        threshold = threshold / 100
        binary_pred = [1 if pred >= threshold else 0 for pred in y_pred]
        f1 = metrics.f1_score(y_true, binary_pred)
        # f1 = matthews_corrcoef(y_true, binary_pred)
        if f1 > best_f1:
            best_f1 = f1
            best_threshold = threshold

    for value in y_pred:
        if value < best_threshold:
            y_pred_new.append(0)
        else:
            y_pred_new.append(1)
    y_pred_new = np.array(y_pred_new)

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred_new).ravel()
    fpr, tpr, thresholds = roc_curve(y_true, y_pred_new, pos_label=1)
    roc_auc = metrics.auc(fpr, tpr)
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred_new)
    auprc = metrics.auc(recall, precision)
    thd =best_threshold

    # plot_roc_curve(y_true,y_pred)
    # plt.title("ROC")

    print("Matthews相关系数: " + str(matthews_corrcoef(y_true, y_pred_new)))
    print("ACC: ", (tp + tn) / (tp + tn + fp + fn))
    print("AUC: ", roc_auc)
    print('sensitivity/recall:', tp / (tp + fn))
    print('specificity:', tn / (tn + fp))
    print('precision:', tp / (tp + fp))
    print('negative predictive value:', tn / (tn + fn))
    print("F1值: " + str(f1_score(y_true, y_pred_new)))
    print('error rate:', fp / (tp + tn + fp + fn))
    print('false positive rate:', fp / (tn + fp))
    print('false negative rate:', fn / (tp + fn))
    print('TN:', tn, 'FP:', fp, 'FN:', fn, 'TP:', tp)
    print('AUPRC: ' + str(auprc))
    print('best_threshold: ' + str(best_threshold))

    mcc = float(format((matthews_corrcoef(y_true, y_pred_new)), '.4f'))
    acc = float(format((tp + tn) / (tp + tn + fp + fn), '.4f'))
    auc = float(format(roc_auc, '.4f'))
    sen = float(format(tp / (tp + fn), '.4f'))
    spe = float(format(tn / (tn + fp), '.4f'))
    pre = float(format(tp / (tp + fp), '.4f'))

    npv = float(format(tn / (tn + fn), '.4f'))
    f1 = float(format(f1_score(y_true, y_pred_new), '.4f'))
    fpr = float(format(fp / (tn + fp), '.4f'))
    fnr = float(format(fn / (tp + fn), '.4f'))
    auprc = float(format(auprc, '.4f'))

    # df = pd.DataFrame({
    #     'True': y_true,
    #     'Predict_value': y_pred,
    #     'Predict': y_pred_new
    # })

    # try:
    #     # 检查文件是否存在
    #     book = load_workbook(p_result)
    #     # 如果文件存在，删除工作表并保存新文件
    #     book.remove(book.active)
    # except FileNotFoundError:
    #     pass  # 文件不存在，则继续保存新文件
    # # 将 DataFrame 保存到 Excel 文件
    # df.to_excel(p_result, index=False)

    # 保存每一次跑的结果到excel表格
    result = mcc, acc, auc, sen, spe, pre, npv, f1, fpr, fnr, tn, fp, fn, tp, auprc ,thd
    op_toexcel(result, filename)
